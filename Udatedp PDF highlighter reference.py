import re
import os
import csv
import sys
import json
import time
import shutil
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, colorchooser

# ── PDF library ──────────────────────────────────────────────────────
try:
    import fitz  # PyMuPDF
except ImportError:
    messagebox.showerror(
        "Missing Library",
        "PyMuPDF is required. Install it with:\npip install PyMuPDF"
    )
    sys.exit(1)

# ── Optional Excel library ───────────────────────────────────────────
try:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Pre-compiled regex for stripping non-word characters ─────────────
_STRIP_NON_WORD = re.compile(r'[^\w]')


# ═══════════════════════════════════════════════════════════════════════
# PALETTES
# ═══════════════════════════════════════════════════════════════════════
PALETTES = {
    "Whispering Sky Meadows": [
        "#f1f2f3", "#c7dcfa", "#d1d5c3", "#b6d1f7", "#dad17c"
    ],
    "90s Cool Minimalism": [
        "#a3c4d7", "#b8e2e5", "#d4f3f7", "#f0b2d3", "#f6a2b8"
    ],
    "Blue Clouds": [
        "#a4c8df", "#f0f2f4", "#e3e9f2", "#d2e2f9", "#b0c8e8"
    ],
    "Sunset Coral": [
        "#ff6e61", "#f7c5a1", "#f5e0a3", "#a3d55d", "#79c8c8"
    ],
}


# ═══════════════════════════════════════════════════════════════════════
# CONFIG / PERSISTENCE
# ═══════════════════════════════════════════════════════════════════════
def get_default_base_dir():
    docs = os.path.join(os.path.expanduser("~"), "Documents")
    return os.path.join(docs, "PDFHighlighter")


def get_settings_path():
    if getattr(sys, 'frozen', False):
        app_dir = os.path.dirname(sys.executable)
    elif '__file__' in globals():
        app_dir = os.path.dirname(os.path.abspath(__file__))
    else:
        app_dir = os.getcwd()
    return os.path.join(app_dir, "pdf_highlighter_settings.json")


def load_settings():
    path = get_settings_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_settings(settings):
    path = get_settings_path()
    try:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(settings, f, indent=2, ensure_ascii=False)
    except OSError:
        pass


def get_base_dir():
    settings = load_settings()
    return settings.get("base_dir", get_default_base_dir())


def ensure_folders(base_dir):
    input_dir = os.path.join(base_dir, "Input Folder")
    output_dir = os.path.join(base_dir, "Output Folder")
    for folder in (base_dir, input_dir, output_dir):
        os.makedirs(folder, exist_ok=True)
    return input_dir, output_dir


def config_file_path(base_dir):
    return os.path.join(base_dir, "categories.json")


def load_categories(base_dir) -> dict:
    cf = config_file_path(base_dir)
    if os.path.exists(cf):
        try:
            with open(cf, "r", encoding="utf-8") as f:
                return json.load(f)
        except (json.JSONDecodeError, OSError):
            return {}
    return {}


def save_categories(cats: dict, base_dir: str):
    cf = config_file_path(base_dir)
    try:
        with open(cf, "w", encoding="utf-8") as f:
            json.dump(cats, f, indent=2, ensure_ascii=False)
    except OSError:
        pass


# ═══════════════════════════════════════════════════════════════════════
# HELPER
# ═══════════════════════════════════════════════════════════════════════
def hex_to_rgb(hex_str):
    hex_str = hex_str.lstrip("#")
    return [int(hex_str[i:i + 2], 16) / 255.0 for i in (0, 2, 4)]


def resolve_color(color):
    if isinstance(color, str):
        return hex_to_rgb(color)
    return list(color)


# ═══════════════════════════════════════════════════════════════════════
# HIGHLIGHTING ENGINE
# ═══════════════════════════════════════════════════════════════════════
def highlight_pdf(input_path, output_path, categories, log_func=None,
                  stop_at_references=False):                        # ← NEW
    try:
        doc = fitz.open(input_path)
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Cannot open: {e}")
        return 0, set(), {}

    if log_func:
        log_func(f"  Processing: {os.path.basename(input_path)}")

    total_highlights = 0
    triggered = set()
    matched_keywords = {}

    # ── Pattern to detect a "References" heading ──────────────────────
    # Matches lines that are just "References", "Bibliography", etc.
    _REF_HEADING = re.compile(
        r"^\s*(references|bibliography|works\s+cited|literature\s+cited)\s*[^a-zA-Z]{0,5}$",
        re.IGNORECASE
    )

    # Prepare keywords once
    prepared = []
    for cat_name, cat_data in categories.items():
        color = resolve_color(cat_data.get("color", [1, 1, 0]))
        matched_keywords[cat_name] = set()
        for word in cat_data.get("words", []):
            parts = [
                _STRIP_NON_WORD.sub('', p).lower()
                for p in word.split()
                if _STRIP_NON_WORD.sub('', p)
            ]
            if parts:
                prepared.append((cat_name, color, word, parts))

    ref_found = False                                               # ← NEW

    for page_num in range(len(doc)):
        page = doc[page_num]
        word_list = page.get_text("words")

        if not word_list:
            continue

        # ── Stop-at-references: detect heading & trim word_list ───
        if stop_at_references and not ref_found:                    # ← NEW
            # Group words into lines by (block_no, line_no)
            from collections import defaultdict
            lines = defaultdict(list)
            for idx, w in enumerate(word_list):
                # w = (x0, y0, x1, y1, text, block_no, line_no, word_no)
                key = (w[5], w[6])
                lines[key].append((idx, w))

            cutoff_index = None
            for key in sorted(lines.keys()):
                line_text = " ".join(w[4] for _, w in lines[key])
                if _REF_HEADING.match(line_text):
                    # Cut at the first word of this line
                    cutoff_index = lines[key][0][0]
                    ref_found = True
                    if log_func:
                        log_func(f"    ⊘ References heading found on page {page_num + 1}, stopping highlights")
                    break

            if cutoff_index is not None:
                word_list = word_list[:cutoff_index]
                if not word_list:
                    continue
        elif stop_at_references and ref_found:                      # ← NEW
            # All subsequent pages are skipped
            continue

        cleaned_words = [
            _STRIP_NON_WORD.sub('', w[4]).lower()
            for w in word_list
        ]
        num_words = len(cleaned_words)

        for cat_name, color, kw_original, kw_parts in prepared:
            kw_len = len(kw_parts)
            if kw_len > num_words:
                continue

            for i in range(num_words - kw_len + 1):
                if cleaned_words[i] != kw_parts[0]:
                    continue

                match = True
                for j in range(1, kw_len):
                    if cleaned_words[i + j] != kw_parts[j]:
                        match = False
                        break

                if match:
                    for j in range(kw_len):
                        w = word_list[i + j]
                        rect = fitz.Rect(w[0], w[1], w[2], w[3])
                        try:
                            annot = page.add_highlight_annot(rect)
                            annot.set_colors(stroke=color)
                            annot.update()
                        except Exception:
                            pass

                    total_highlights += 1
                    triggered.add(cat_name)
                    matched_keywords[cat_name].add(kw_original)

    try:
        doc.save(output_path)
        if log_func:
            log_func(f"  ✓ Saved: {os.path.basename(output_path)}")
    except Exception as e:
        if log_func:
            log_func(f"  ✗ Could not save: {e}")
    finally:
        doc.close()

    return total_highlights, triggered, matched_keywords



# ═══════════════════════════════════════════════════════════════════════
# REPORT HELPERS
# ═══════════════════════════════════════════════════════════════════════
def _scan_pdfs_from_results(all_results, filtered_cats):
    rows = []
    for filename, matched_keywords in all_results.items():
        row = {"Filename": filename}
        for cat_name in filtered_cats:
            cat_words = set(
                w.lower() for w in filtered_cats[cat_name].get("words", [])
            )
            matched = [
                w for w in matched_keywords.get(cat_name, set())
                if w.lower() in cat_words
            ]
            row[cat_name] = matched
        rows.append(row)
    return rows


def _scan_pdfs_fallback(input_dir, filtered_cats):
    pdf_files = sorted([
        f for f in os.listdir(input_dir)
        if f.lower().endswith(".pdf")
    ])
    rows = []

    keyword_patterns = {}
    for cat_name, cat_data in filtered_cats.items():
        patterns = []
        for word in cat_data["words"]:
            pattern = re.compile(
                r'(?<!\w)' + re.escape(word.lower()) + r'(?!\w)'
            )
            patterns.append((word, pattern))
        keyword_patterns[cat_name] = patterns

    for filename in pdf_files:
        filepath = os.path.join(input_dir, filename)
        try:
            doc = fitz.open(filepath)
        except Exception:
            continue

        try:
            text_parts = []
            for page_num in range(len(doc)):
                text_parts.append(
                    doc[page_num].get_text("text").lower()
                )
            full_text = " ".join(text_parts)
        finally:
            doc.close()

        row = {"Filename": filename}
        for cat_name, patterns in keyword_patterns.items():
            matched = [
                word for word, pattern in patterns
                if pattern.search(full_text)
            ]
            row[cat_name] = matched
        rows.append(row)

    return rows


def generate_csv_report(input_dir, output_dir, categories, log_func=None,
                        report_categories=None, all_results=None):
    start = time.time()
    filtered_cats = {
        k: v for k, v in categories.items()
        if report_categories is None or k in report_categories
    }
    cat_names = list(filtered_cats.keys())

    if all_results is not None:
        rows = _scan_pdfs_from_results(all_results, filtered_cats)
    else:
        rows = _scan_pdfs_fallback(input_dir, filtered_cats)

    report_path = os.path.join(output_dir, "report.csv")

    try:
        with open(report_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(["#", "Filename"] + cat_names)
            for idx, row in enumerate(rows, start=1):
                csv_row = [idx, row["Filename"]]
                for cat in cat_names:
                    csv_row.append(", ".join(row[cat]) if row[cat] else "")
                writer.writerow(csv_row)
    except OSError as e:
        if log_func:
            log_func(f"  ✗ Could not write CSV: {e}")
        return None

    elapsed = time.time() - start
    if log_func:
        log_func(f"  ✓ CSV report saved ({elapsed:.1f}s)")
    return report_path


def generate_excel_report(input_dir, output_dir, categories, log_func=None,
                          report_categories=None, all_results=None):
    if not HAS_OPENPYXL:
        if log_func:
            log_func("⚠ openpyxl not installed — skipping Excel report.")
        return None

    start = time.time()
    filtered_cats = {
        k: v for k, v in categories.items()
        if report_categories is None or k in report_categories
    }
    cat_names = list(filtered_cats.keys())

    if all_results is not None:
        rows = _scan_pdfs_from_results(all_results, filtered_cats)
    else:
        rows = _scan_pdfs_fallback(input_dir, filtered_cats)

    GREEN_FILL = PatternFill(
        start_color="E7FEEF", end_color="E7FEEF", fill_type="solid"
    )
    RED_FILL = PatternFill(
        start_color="FA8F8F", end_color="FA8F8F", fill_type="solid"
    )

    wb = Workbook()

    def _write_sheet(ws, binary=False):
        cell_num = ws.cell(row=1, column=1, value="#")
        cell_num.font = Font(bold=True)
        cell_num.alignment = Alignment(horizontal="center")

        cell_fn = ws.cell(row=1, column=2, value="Filename")
        cell_fn.font = Font(bold=True)
        cell_fn.alignment = Alignment(horizontal="center")

        for col_idx, cat_name in enumerate(cat_names, start=3):
            cell = ws.cell(row=1, column=col_idx, value=cat_name)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            rgb = resolve_color(filtered_cats[cat_name]["color"])
            hex_color = "{:02X}{:02X}{:02X}".format(
                int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255)
            )
            cell.fill = PatternFill(
                start_color=hex_color, end_color=hex_color, fill_type="solid"
            )

        for row_idx, row in enumerate(rows, start=2):
            ws.cell(row=row_idx, column=1, value=row_idx - 1)
            ws.cell(row=row_idx, column=2, value=row["Filename"])
            for col_idx, cat in enumerate(cat_names, start=3):
                if binary:
                    val = 1 if row[cat] else 0
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.fill = GREEN_FILL if val == 1 else RED_FILL
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell = ws.cell(
                        row=row_idx, column=col_idx,
                        value=", ".join(row[cat]) if row[cat] else ""
                    )

    ws_binary = wb.active
    ws_binary.title = "Binary"
    _write_sheet(ws_binary, binary=True)

    ws_detail = wb.create_sheet("Details")
    _write_sheet(ws_detail, binary=False)

    report_path = os.path.join(output_dir, "report.xlsx")
    try:
        wb.save(report_path)
    except OSError as e:
        if log_func:
            log_func(f"  ✗ Could not write Excel: {e}")
        return None

    elapsed = time.time() - start
    if log_func:
        log_func(f"  ✓ Excel report saved ({elapsed:.1f}s)")
    return report_path


# ═══════════════════════════════════════════════════════════════════════
# MAIN PROCESSING
# ═══════════════════════════════════════════════════════════════════════
def process_all_pdfs(input_dir, output_dir, categories, log_func=None,
                     done_callback=None, report_categories=None,
                     do_csv=False, do_excel=False,
                     result_callback=None,
                     stop_at_references=False):          # ← NEW
    pdf_files = sorted([
        f for f in os.listdir(input_dir)
        if f.lower().endswith(".pdf")
    ])

    if not pdf_files:
        if log_func:
            log_func("No PDF files found in Input Folder.")
        if done_callback:
            done_callback()
        return

    if log_func:
        log_func(f"Processing {len(pdf_files)} PDF(s)...\n")

    total = 0
    all_triggered = set()
    all_results = {}
    start = time.time()

    for filename in pdf_files:
        input_path = os.path.join(input_dir, filename)
        output_path = os.path.join(output_dir, filename)
        try:
            count, triggered, matched_kw = highlight_pdf(
                input_path, output_path, categories, log_func,
                stop_at_references=stop_at_references    # ← NEW
            )
            total += count
            all_triggered |= triggered
            all_results[filename] = matched_kw
        except Exception as e:
            if log_func:
                log_func(f"  ✖ {filename}: ERROR — {e}")

    elapsed = time.time() - start

    if log_func:
        log_func(f"\n{'═' * 40}")
        log_func(
            f"Done! {total} total highlights across "
            f"{len(pdf_files)} file(s) in {elapsed:.1f}s"
        )
        if all_triggered:
            log_func(
                f"Categories triggered: "
                f"{', '.join(sorted(all_triggered))}"
            )
        log_func(f"Output folder: {output_dir}")

    if report_categories is not None:
        if do_csv:
            generate_csv_report(
                input_dir, output_dir, categories, log_func,
                report_categories, all_results=all_results
            )
        if do_excel:
            generate_excel_report(
                input_dir, output_dir, categories, log_func,
                report_categories, all_results=all_results
            )

    if result_callback:
        result_callback(all_results, categories)

    if done_callback:
        done_callback()



# ═══════════════════════════════════════════════════════════════════════
# REPORT WINDOW — Interactive post-scan file organizer
# ═══════════════════════════════════════════════════════════════════════
class ReportWindow(tk.Toplevel):
    """
    Shows scan results in a table. User can:
    - View binary (1/0) or toggle to see matched keywords
    - Sort by any column
    - Search/filter by filename
    - Set category rules (must be true / must be false)
    - Create folders with copies of matching PDFs
    """

    def __init__(self, parent, all_results, categories, output_dir):
        super().__init__(parent)
        self.title("Scan Report — File Organizer")
        self.geometry("1000x700")
        self.minsize(800, 500)

        self.all_results = all_results  # {filename: {cat: set(keywords)}}
        self.categories = categories
        self.output_dir = output_dir
        self.cat_names = list(categories.keys())

        # Build row data: list of dicts {Filename, cat1: [...], cat2: [...]}
        self.rows = self._build_rows()

        # Sort state
        self._sort_col = None
        self._sort_reverse = False

        # Show detail mode
        self._show_detail = tk.BooleanVar(value=False)

        # Rule variables: each category can be "Any", "Must be 1", "Must be 0"
        self.rule_vars = {}

        self._build_ui()
        self._populate_table()

    def _build_rows(self):
        rows = []
        for filename, matched_keywords in self.all_results.items():
            row = {"Filename": filename}
            for cat_name in self.cat_names:
                cat_words = set(
                    w.lower()
                    for w in self.categories[cat_name].get("words", [])
                )
                matched = [
                    w for w in matched_keywords.get(cat_name, set())
                    if w.lower() in cat_words
                ]
                row[cat_name] = matched
            rows.append(row)
        return rows

    def _build_ui(self):
        # ── Top toolbar ──
        toolbar = ttk.Frame(self)
        toolbar.pack(fill="x", padx=10, pady=(10, 5))

        ttk.Label(toolbar, text="Search filename:").pack(side="left")
        self.search_var = tk.StringVar()
        self.search_var.trace_add("write", lambda *_: self._populate_table())
        search_entry = ttk.Entry(
            toolbar, textvariable=self.search_var, width=25
        )
        search_entry.pack(side="left", padx=(5, 15))

        ttk.Checkbutton(
            toolbar, text="Show matched keywords",
            variable=self._show_detail,
            command=self._populate_table
        ).pack(side="left", padx=5)

        ttk.Label(
            toolbar,
            text=f"{len(self.rows)} files  |  {len(self.cat_names)} categories",
            foreground="gray"
        ).pack(side="right")

        # ── Treeview table ──
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=10, pady=5)

        columns = ["#", "Filename"] + self.cat_names
        self.tree = ttk.Treeview(
            tree_frame, columns=columns, show="headings",
            selectmode="extended"
        )

        # Scrollbars
        vsb = ttk.Scrollbar(
            tree_frame, orient="vertical", command=self.tree.yview
        )
        hsb = ttk.Scrollbar(
            tree_frame, orient="horizontal", command=self.tree.xview
        )
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)

        # Column headings
        self.tree.heading("#", text="#", command=lambda: self._sort_by("#"))
        self.tree.column("#", width=40, stretch=False, anchor="center")

        self.tree.heading(
            "Filename", text="Filename",
            command=lambda: self._sort_by("Filename")
        )
        self.tree.column("Filename", width=250, anchor="w")

        for cat in self.cat_names:
            self.tree.heading(
                cat, text=cat,
                command=lambda c=cat: self._sort_by(c)
            )
            self.tree.column(cat, width=80, anchor="center")

        # Tag colors for binary display
        self.tree.tag_configure("hit", background="#E7FEEF")
        self.tree.tag_configure("miss", background="#FAEAEA")

        # ── Rules frame ──
        rules_outer = ttk.LabelFrame(self, text="Filter Rules")
        rules_outer.pack(fill="x", padx=10, pady=5)

        rules_scroll = ttk.Frame(rules_outer)
        rules_scroll.pack(fill="x", padx=5, pady=5)

        # Create a rule dropdown for each category
        for idx, cat in enumerate(self.cat_names):
            col = idx % 4
            row = idx // 4

            frame = ttk.Frame(rules_scroll)
            frame.grid(row=row, column=col, padx=5, pady=2, sticky="w")

            ttk.Label(frame, text=f"{cat}:", width=15, anchor="w").pack(
                side="left"
            )
            var = tk.StringVar(value="Any")
            self.rule_vars[cat] = var
            combo = ttk.Combobox(
                frame, textvariable=var,
                values=["Any", "Must be 1", "Must be 0"],
                state="readonly", width=12
            )
            combo.pack(side="left")
            combo.bind(
                "<<ComboboxSelected>>", lambda e: self._populate_table()
            )

        # ── Bottom action bar ──
        action_frame = ttk.Frame(self)
        action_frame.pack(fill="x", padx=10, pady=(5, 10))

        self.match_label = ttk.Label(
            action_frame, text="Matching: 0 / 0", font=("", 10, "bold")
        )
        self.match_label.pack(side="left", padx=5)

        ttk.Button(
            action_frame, text="Reset Rules",
            command=self._reset_rules
        ).pack(side="left", padx=10)

        ttk.Button(
            action_frame, text="📁 Create Folder from Filter",
            command=self._create_filtered_folder
        ).pack(side="right", padx=5)

        ttk.Button(
            action_frame, text="Open Output Folder",
            command=self._open_output
        ).pack(side="right", padx=5)

    def _get_filtered_rows(self):
        """Apply search + rules, return list of (index, row) tuples."""
        search = self.search_var.get().strip().lower()
        filtered = []

        for idx, row in enumerate(self.rows):
            # Search filter
            if search and search not in row["Filename"].lower():
                continue

            # Rule filters
            passes = True
            for cat, var in self.rule_vars.items():
                rule = var.get()
                has_match = bool(row.get(cat))
                if rule == "Must be 1" and not has_match:
                    passes = False
                    break
                elif rule == "Must be 0" and has_match:
                    passes = False
                    break

            if passes:
                filtered.append((idx, row))

        return filtered

    def _populate_table(self):
        """Refresh the treeview with current filters applied."""
        self.tree.delete(*self.tree.get_children())

        filtered = self._get_filtered_rows()
        show_detail = self._show_detail.get()

        for display_idx, (orig_idx, row) in enumerate(filtered, start=1):
            values = [display_idx, row["Filename"]]

            for cat in self.cat_names:
                matched = row.get(cat, [])
                if show_detail:
                    values.append(", ".join(matched) if matched else "—")
                else:
                    values.append("1" if matched else "0")

            self.tree.insert("", "end", values=values)

        total = len(self.rows)
        shown = len(filtered)
        self.match_label.configure(text=f"Matching: {shown} / {total}")

    def _sort_by(self, col):
        """Sort table by clicking column header."""
        if self._sort_col == col:
            self._sort_reverse = not self._sort_reverse
        else:
            self._sort_col = col
            self._sort_reverse = False

        if col == "#":
            # Sort by original order — just repopulate
            self._sort_col = None
            self._populate_table()
            return

        if col == "Filename":
            self.rows.sort(
                key=lambda r: r["Filename"].lower(),
                reverse=self._sort_reverse
            )
        else:
            # Sort by number of matches in that category
            self.rows.sort(
                key=lambda r: len(r.get(col, [])),
                reverse=self._sort_reverse
            )

        self._populate_table()

    def _reset_rules(self):
        for var in self.rule_vars.values():
            var.set("Any")
        self.search_var.set("")
        self._populate_table()

    def _create_filtered_folder(self):
        """Copy matching PDFs into a named subfolder."""
        filtered = self._get_filtered_rows()

        if not filtered:
            messagebox.showinfo(
                "No matches",
                "No files match the current filter rules.",
                parent=self
            )
            return

        # Build a folder name from the active rules
        rule_parts = []
        for cat, var in self.rule_vars.items():
            rule = var.get()
            if rule == "Must be 1":
                rule_parts.append(f"{cat}=YES")
            elif rule == "Must be 0":
                rule_parts.append(f"{cat}=NO")

        if rule_parts:
            folder_name = "Filter_" + "_".join(rule_parts)
        else:
            folder_name = "Filter_AllFiles"

        # Clean folder name of invalid characters
        folder_name = re.sub(r'[<>:"/\\|?*]', '_', folder_name)

        # Limit length to avoid OS issues
        if len(folder_name) > 100:
            folder_name = folder_name[:100]

        target_dir = os.path.join(self.output_dir, folder_name)

        # If folder exists, ask what to do
        if os.path.exists(target_dir):
            answer = messagebox.askyesnocancel(
                "Folder exists",
                f"'{folder_name}' already exists.\n\n"
                "Yes = Overwrite contents\n"
                "No = Add a number suffix\n"
                "Cancel = Abort",
                parent=self
            )
            if answer is None:
                return
            elif answer:
                # Overwrite — clear the folder
                try:
                    shutil.rmtree(target_dir)
                except OSError as e:
                    messagebox.showerror(
                        "Error", f"Could not remove folder: {e}",
                        parent=self
                    )
                    return
            else:
                # Add suffix
                counter = 2
                while os.path.exists(f"{target_dir}_{counter}"):
                    counter += 1
                target_dir = f"{target_dir}_{counter}"
                folder_name = os.path.basename(target_dir)

        try:
            os.makedirs(target_dir, exist_ok=True)
        except OSError as e:
            messagebox.showerror(
                "Error", f"Could not create folder: {e}", parent=self
            )
            return

        # Copy files
        copied = 0
        errors = 0
        for _, row in filtered:
            src = os.path.join(self.output_dir, row["Filename"])
            dst = os.path.join(target_dir, row["Filename"])
            try:
                if os.path.exists(src):
                    shutil.copy2(src, dst)
                    copied += 1
                else:
                    errors += 1
            except OSError:
                errors += 1

        msg = f"✓ Copied {copied} file(s) to:\n{target_dir}"
        if errors:
            msg += f"\n\n⚠ {errors} file(s) could not be copied."

        messagebox.showinfo("Folder Created", msg, parent=self)

    def _open_output(self):
        if os.path.isdir(self.output_dir):
            try:
                if sys.platform == "win32":
                    os.startfile(self.output_dir)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", self.output_dir])
                else:
                    subprocess.Popen(["xdg-open", self.output_dir])
            except OSError:
                pass


# ═══════════════════════════════════════════════════════════════════════
# GUI
# ═══════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PDF Keyword Highlighter")
        self.geometry("750x650")
        self.resizable(True, True)

        self.base_dir = get_base_dir()
        self.input_dir, self.output_dir = ensure_folders(self.base_dir)
        self.categories = load_categories(self.base_dir)
        self.report_csv = tk.BooleanVar(value=False)
        self.report_excel = tk.BooleanVar(value=False)
        self.stop_at_refs = tk.BooleanVar(value=False) 
        self.report_cat_vars = {}
        self.cat_active_vars = {}  # NEW: which categories are enabled for run
        self._is_running = False

        self._build_ui()
        self._refresh_cat_list()
        self._update_file_count()

        self.protocol("WM_DELETE_WINDOW", self._on_close)

    def _on_close(self):
        if self._is_running:
            if not messagebox.askyesno(
                "Processing",
                "Processing is still running. Close anyway?"
            ):
                return
        self.destroy()

    def _build_ui(self):
        # ── Top frame: working directory ──
        dir_frame = ttk.LabelFrame(self, text="Working Directory")
        dir_frame.pack(fill="x", padx=10, pady=(10, 5))

        self.dir_label = ttk.Label(
            dir_frame, text=self.base_dir,
            foreground="blue", cursor="hand2"
        )
        self.dir_label.pack(
            side="left", padx=10, pady=5, fill="x", expand=True
        )
        self.dir_label.bind(
            "<Button-1>", lambda e: self._open_folder(self.base_dir)
        )

        ttk.Button(
            dir_frame, text="Change…",
            command=self._change_base_dir
        ).pack(side="right", padx=10, pady=5)

        # ── Category frame with checkboxes ──
        cat_frame = ttk.LabelFrame(self, text="Categories (✓ = active for run)")
        cat_frame.pack(fill="both", padx=10, pady=5, expand=True)

        # Scrollable frame for categories with checkboxes
        cat_canvas = tk.Canvas(cat_frame, highlightthickness=0)
        cat_scrollbar = ttk.Scrollbar(
            cat_frame, orient="vertical", command=cat_canvas.yview
        )
        self.cat_inner_frame = ttk.Frame(cat_canvas)

        self.cat_inner_frame.bind(
            "<Configure>",
            lambda e: cat_canvas.configure(
                scrollregion=cat_canvas.bbox("all")
            )
        )

        cat_canvas.create_window(
            (0, 0), window=self.cat_inner_frame, anchor="nw"
        )
        cat_canvas.configure(yscrollcommand=cat_scrollbar.set)

        cat_canvas.pack(side="left", fill="both", expand=True, padx=5, pady=5)
        cat_scrollbar.pack(side="right", fill="y", pady=5)

        # Store canvas reference for mouse wheel binding
        self._cat_canvas = cat_canvas

        # Bind mouse wheel to scroll
        def _on_mousewheel(event):
            cat_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        cat_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        btn_row = ttk.Frame(cat_frame)
        btn_row.pack(fill="x", padx=10, pady=(0, 5))
        ttk.Button(
            btn_row, text="Add", command=self._add_cat
        ).pack(side="left", padx=2)
        ttk.Button(
            btn_row, text="Edit", command=self._edit_cat
        ).pack(side="left", padx=2)
        ttk.Button(
            btn_row, text="Delete", command=self._delete_cat
        ).pack(side="left", padx=2)
        ttk.Separator(btn_row, orient="vertical").pack(
            side="left", padx=8, fill="y"
        )
        ttk.Button(
            btn_row, text="Select All",
            command=lambda: self._set_all_cats(True)
        ).pack(side="left", padx=2)
        ttk.Button(
            btn_row, text="Deselect All",
            command=lambda: self._set_all_cats(False)
        ).pack(side="left", padx=2)

        # ── Report options ──
        report_frame = ttk.LabelFrame(self, text="Report Options")
        report_frame.pack(fill="x", padx=10, pady=5)

        ttk.Checkbutton(
            report_frame, text="CSV Report",
            variable=self.report_csv
        ).pack(side="left", padx=10, pady=5)
        ttk.Checkbutton(
            report_frame, text="Excel Report",
            variable=self.report_excel
        ).pack(side="left", padx=5, pady=5)
        ttk.Button(
            report_frame, text="Select Categories…",
            command=self._select_report_categories
        ).pack(side="left", padx=10, pady=5)

        # ── Processing Options ──
        options_frame = ttk.LabelFrame(self, text="Processing Options")
        options_frame.pack(fill="x", padx=10, pady=5)

        ttk.Checkbutton(
            options_frame,
            text="Stop highlighting at References section",
            variable=self.stop_at_refs
        ).pack(anchor="w", padx=10, pady=5)

        # ── File count + Run ──
        run_frame = ttk.Frame(self)
        run_frame.pack(fill="x", padx=10, pady=5)

        self.file_count_label = ttk.Label(
            run_frame, text="PDFs in input: 0"
        )
        self.file_count_label.pack(side="left", padx=5)

        ttk.Button(
            run_frame, text="↻ Refresh",
            command=self._update_file_count
        ).pack(side="left", padx=5)
        ttk.Button(
            run_frame, text="Open Input Folder",
            command=lambda: self._open_folder(self.input_dir)
        ).pack(side="left", padx=5)
        ttk.Button(
            run_frame, text="Open Output Folder",
            command=lambda: self._open_folder(self.output_dir)
        ).pack(side="left", padx=5)

        ttk.Button(
            run_frame, text="📂 Import Report",
            command=self._import_report
        ).pack(side="left", padx=5)

        self.run_btn = ttk.Button(
            run_frame, text="▶  Run Highlighter",
            command=self._run
        )
        self.run_btn.pack(side="right", padx=5)

        # ── Log ──
        log_frame = ttk.LabelFrame(self, text="Log")
        log_frame.pack(fill="both", padx=10, pady=(5, 10), expand=True)

        self.log_text = tk.Text(
            log_frame, height=10, state="disabled",
            wrap="word", bg="#1e1e1e", fg="#d4d4d4",
            font=("Consolas", 10)
        )
        self.log_text.pack(fill="both", padx=5, pady=5, expand=True)

    # ── Category list with checkboxes ──

    def _refresh_cat_list(self):
        """Rebuild the category list with checkboxes."""
        for widget in self.cat_inner_frame.winfo_children():
            widget.destroy()

        self.cat_active_vars = {}
        self._cat_labels = {}

        for idx, (name, info) in enumerate(self.categories.items()):
            color = resolve_color(info.get("color", [1, 1, 0]))
            n_words = len(info.get("words", []))
            hex_c = self._rgb_to_hex(color)

            row_frame = ttk.Frame(self.cat_inner_frame)
            row_frame.pack(fill="x", padx=5, pady=1)

            # Checkbox for active/inactive
            var = tk.BooleanVar(value=True)
            self.cat_active_vars[name] = var
            cb = ttk.Checkbutton(row_frame, variable=var)
            cb.pack(side="left")

            # Color swatch
            swatch = tk.Label(
                row_frame, bg=hex_c, width=2, height=1,
                relief="solid", borderwidth=1
            )
            swatch.pack(side="left", padx=(2, 5))

            # Category name and info
            label = ttk.Label(
                row_frame,
                text=f"{name}  ({n_words} keywords)  [{hex_c}]",
                cursor="hand2"
            )
            label.pack(side="left", fill="x", expand=True)

            # Double-click to edit
            label.bind(
                "<Double-1>",
                lambda e, n=name: self._edit_cat(preselect=n)
            )

            self._cat_labels[name] = label

    def _set_all_cats(self, state):
        for var in self.cat_active_vars.values():
            var.set(state)

    def _get_active_categories(self):
        """Return dict of only the checked categories."""
        return {
            name: self.categories[name]
            for name, var in self.cat_active_vars.items()
            if var.get() and name in self.categories
        }

    def _rgb_to_hex(self, rgb):
        return "#{:02x}{:02x}{:02x}".format(
            int(rgb[0] * 255), int(rgb[1] * 255), int(rgb[2] * 255)
        )

    def _update_file_count(self):
        try:
            count = len([
                f for f in os.listdir(self.input_dir)
                if f.lower().endswith(".pdf")
            ])
        except FileNotFoundError:
            count = 0
        self.file_count_label.configure(text=f"PDFs in input: {count}")

    def _open_folder(self, path):
        if os.path.isdir(path):
            try:
                if sys.platform == "win32":
                    os.startfile(path)
                elif sys.platform == "darwin":
                    subprocess.Popen(["open", path])
                else:
                    subprocess.Popen(["xdg-open", path])
            except OSError:
                pass

    def _change_base_dir(self):
        new_dir = filedialog.askdirectory(
            title="Select working directory",
            initialdir=self.base_dir
        )
        if new_dir:
            self.base_dir = new_dir
            self.input_dir, self.output_dir = ensure_folders(self.base_dir)
            self.categories = load_categories(self.base_dir)
            self.dir_label.configure(text=self.base_dir)
            settings = load_settings()
            settings["base_dir"] = self.base_dir
            save_settings(settings)
            self._refresh_cat_list()
            self._update_file_count()

    def _log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert("end", msg + "\n")
        self.log_text.see("end")
        self.log_text.configure(state="disabled")

    # ── Category CRUD ──

    def _add_cat(self):
        self._cat_dialog()

    def _edit_cat(self, preselect=None):
        if preselect:
            name = preselect
        else:
            # Find which category label area might be focused
            # Fall back to a selection dialog
            names = list(self.categories.keys())
            if not names:
                return
            if len(names) == 1:
                name = names[0]
            else:
                name = self._pick_category_dialog("Edit Category")
                if not name:
                    return

        if name in self.categories:
            self._cat_dialog(existing_name=name)

    def _pick_category_dialog(self, title):
        """Simple dialog to pick a category by name."""
        dlg = tk.Toplevel(self)
        dlg.title(title)
        dlg.geometry("300x350")
        dlg.grab_set()

        result = {"name": None}

        ttk.Label(dlg, text="Select a category:").pack(
            anchor="w", padx=10, pady=10
        )

        listbox = tk.Listbox(dlg, height=10)
        listbox.pack(fill="both", expand=True, padx=10, pady=5)

        for name in self.categories:
            listbox.insert("end", name)

        def on_ok():
            sel = listbox.curselection()
            if sel:
                result["name"] = listbox.get(sel[0])
            dlg.destroy()

        ttk.Button(dlg, text="OK", command=on_ok).pack(pady=10)

        dlg.wait_window()
        return result["name"]

    def _delete_cat(self):
        name = self._pick_category_dialog("Delete Category")
        if name and name in self.categories:
            if messagebox.askyesno(
                "Delete", f"Delete category '{name}'?"
            ):
                del self.categories[name]
                save_categories(self.categories, self.base_dir)
                self._refresh_cat_list()

    def _cat_dialog(self, existing_name=None):
        dlg = tk.Toplevel(self)
        dlg.title("Edit Category" if existing_name else "Add Category")
        dlg.geometry("500x550")
        dlg.grab_set()

        existing = self.categories.get(existing_name, {}) if existing_name else {}

        # Name
        ttk.Label(dlg, text="Category Name:").pack(
            anchor="w", padx=10, pady=(10, 0)
        )
        name_var = tk.StringVar(value=existing_name or "")
        ttk.Entry(dlg, textvariable=name_var).pack(
            fill="x", padx=10, pady=5
        )

        # Keywords
        ttk.Label(dlg, text="Keywords (comma-separated):").pack(
            anchor="w", padx=10, pady=(10, 0)
        )
        word_box = tk.Text(dlg, height=6)
        word_box.pack(fill="x", padx=10, pady=5)
        if existing.get("words"):
            word_box.insert("1.0", ", ".join(existing["words"]))

        # Color
        color_holder = {
            "rgb": resolve_color(existing.get("color", [1, 1, 0]))
        }

        color_main = ttk.LabelFrame(dlg, text="Highlight Color")
        color_main.pack(fill="x", padx=10, pady=5)

        swatch_frame = ttk.Frame(color_main)
        swatch_frame.pack(fill="x", padx=10, pady=5)

        ttk.Label(swatch_frame, text="Current:").pack(side="left")
        color_swatch = tk.Label(
            swatch_frame,
            bg=self._rgb_to_hex(color_holder["rgb"]),
            width=6, height=2, relief="solid", borderwidth=1
        )
        color_swatch.pack(side="left", padx=(5, 15))

        hex_var = tk.StringVar(
            value=self._rgb_to_hex(color_holder["rgb"])
        )
        ttk.Label(swatch_frame, text="Hex:").pack(side="left")
        hex_entry = ttk.Entry(swatch_frame, textvariable=hex_var, width=10)
        hex_entry.pack(side="left", padx=2)

        def _apply_color(hex_str):
            try:
                rgb = hex_to_rgb(hex_str)
                color_holder["rgb"] = rgb
                color_swatch.configure(bg=hex_str)
                hex_var.set(hex_str)
            except Exception:
                pass

        def apply_hex():
            _apply_color(hex_var.get().strip())

        ttk.Button(
            swatch_frame, text="Apply", command=apply_hex
        ).pack(side="left")

        def pick_system_color():
            result = colorchooser.askcolor(
                color=self._rgb_to_hex(color_holder["rgb"]),
                title="Pick highlight color", parent=dlg
            )
            if result and result[0]:
                r, g, b = result[0]
                _apply_color(
                    "#{:02x}{:02x}{:02x}".format(int(r), int(g), int(b))
                )

        ttk.Button(
            color_main, text="System Color Picker…",
            command=pick_system_color
        ).pack(anchor="w", padx=10, pady=(0, 5))

        # Palette section
        palette_frame = ttk.LabelFrame(
            color_main, text="Auto-Generate from Palette"
        )
        palette_frame.pack(fill="x", padx=10, pady=(0, 10))

        sel_frame = ttk.Frame(palette_frame)
        sel_frame.pack(fill="x", padx=5, pady=5)

        palette_var = tk.StringVar(value=list(PALETTES.keys())[0])
        ttk.Label(sel_frame, text="Palette:").pack(side="left")
        palette_combo = ttk.Combobox(
            sel_frame, textvariable=palette_var,
            values=list(PALETTES.keys()),
            state="readonly", width=24
        )
        palette_combo.pack(side="left", padx=5)

        preview_frame = ttk.Frame(palette_frame)
        preview_frame.pack(fill="x", padx=5, pady=(0, 5))
        swatch_labels = []

        def update_preview(*_):
            for lbl in swatch_labels:
                lbl.destroy()
            swatch_labels.clear()
            pal = PALETTES.get(palette_var.get(), [])
            for i, c in enumerate(pal):
                lbl = tk.Label(
                    preview_frame, bg=c, width=4, height=2,
                    relief="solid", borderwidth=1, cursor="hand2"
                )
                lbl.pack(side="left", padx=2)
                lbl.bind("<Button-1>", lambda e, col=c: _apply_color(col))
                swatch_labels.append(lbl)

        palette_combo.bind("<<ComboboxSelected>>", update_preview)
        update_preview()

        # Save / Cancel
        def save():
            name = name_var.get().strip()
            if not name:
                messagebox.showwarning(
                    "Name", "Enter a category name.", parent=dlg
                )
                return

            raw = word_box.get("1.0", "end").strip()
            words = [w.strip() for w in raw.split(",") if w.strip()]
            if not words:
                messagebox.showwarning(
                    "Keywords", "Enter at least one keyword.", parent=dlg
                )
                return

            if existing_name and name != existing_name:
                if existing_name in self.categories:
                    del self.categories[existing_name]

            self.categories[name] = {
                "words": words,
                "color": color_holder["rgb"]
            }
            save_categories(self.categories, self.base_dir)
            self._refresh_cat_list()
            dlg.destroy()

        btn_frame = ttk.Frame(dlg)
        btn_frame.pack(pady=10)
        ttk.Button(
            btn_frame, text="Save", command=save
        ).pack(side="left", padx=5)
        ttk.Button(
            btn_frame, text="Cancel", command=dlg.destroy
        ).pack(side="left", padx=5)

    def _select_report_categories(self):
        if not self.categories:
            messagebox.showinfo("No categories", "Create categories first.")
            return

        dlg = tk.Toplevel(self)
        dlg.title("Select Report Categories")
        dlg.geometry("300x400")
        dlg.grab_set()

        ttk.Label(dlg, text="Include in report:").pack(
            anchor="w", padx=10, pady=10
        )

        frame = ttk.Frame(dlg)
        frame.pack(fill="both", expand=True, padx=10)

        for name in self.categories:
            if name not in self.report_cat_vars:
                self.report_cat_vars[name] = tk.BooleanVar(value=True)
            ttk.Checkbutton(
                frame, text=name, variable=self.report_cat_vars[name]
            ).pack(anchor="w", pady=2)

        ttk.Button(
            dlg, text="OK", command=dlg.destroy
        ).pack(pady=10)

    # ── Run ──

    def _import_report(self):
        """Import a previously generated Excel report and open the Report Window."""
        if not HAS_OPENPYXL:
            messagebox.showwarning(
                "Missing Library",
                "openpyxl is required to import Excel reports.\n"
                "Install with: pip install openpyxl"
            )
            return

        from openpyxl import load_workbook

        filepath = filedialog.askopenfilename(
            title="Select Excel Report",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=self.output_dir
        )
        if not filepath:
            return

        try:
            wb = load_workbook(filepath, read_only=True)

            if "Detail" in wb.sheetnames:
                ws = wb["Detail"]
            elif "Binary" in wb.sheetnames:
                ws = wb["Binary"]
            else:
                ws = wb.active

            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1)):
                headers.append(cell.value)

            if len(headers) < 3 or headers[0] != "#" or headers[1] != "Filename":
                messagebox.showerror(
                    "Invalid Format",
                    "This doesn't look like a valid report file.\n"
                    "Expected columns: #, Filename, Category1, Category2, ...",
                    parent=self
                )
                wb.close()
                return

            cat_names = headers[2:]

            all_results = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row is None or row[1] is None:
                    continue

                filename = str(row[1]).strip()
                if not filename:
                    continue

                matched_keywords = {}
                for i, cat_name in enumerate(cat_names):
                    col_idx = i + 2
                    cell_value = row[col_idx] if col_idx < len(row) else None

                    if cell_value is None or str(cell_value).strip() == "":
                        matched_keywords[cat_name] = set()
                    elif str(cell_value).strip() in ("0", "1"):
                        if str(cell_value).strip() == "1":
                            matched_keywords[cat_name] = {"(matched)"}
                        else:
                            matched_keywords[cat_name] = set()
                    else:
                        keywords = {
                            kw.strip()
                            for kw in str(cell_value).split(",")
                            if kw.strip()
                        }
                        matched_keywords[cat_name] = keywords

                all_results[filename] = matched_keywords

            wb.close()

            if not all_results:
                messagebox.showinfo(
                    "Empty Report",
                    "No data rows found in the report.",
                    parent=self
                )
                return

            # Collect all keywords seen in the report for each category
            imported_words = {cat: set() for cat in cat_names}
            for filename, matched in all_results.items():
                for cat_name in cat_names:
                    for kw in matched.get(cat_name, set()):
                        imported_words[cat_name].add(kw)

            report_categories = {}
            for cat_name in cat_names:
                if cat_name in self.categories:
                    # Merge: existing words + any from the report
                    existing = self.categories[cat_name].copy()
                    merged_words = set(
                        w.lower() for w in existing.get("words", [])
                    ) | set(w.lower() for w in imported_words[cat_name])
                    existing["words"] = list(merged_words)
                    report_categories[cat_name] = existing
                else:
                    report_categories[cat_name] = {
                        "color": [1, 1, 0],
                        "words": list(imported_words[cat_name])
                    }


            self._log(f"✓ Imported report: {os.path.basename(filepath)}")
            self._log(f"  {len(all_results)} file(s), {len(cat_names)} category(ies)")

            ReportWindow(self, all_results, report_categories, self.output_dir)

        except Exception as e:
            messagebox.showerror(
                "Import Error",
                f"Could not read the report file:\n{e}",
                parent=self
            )

    def _run(self):
        if self._is_running:
            return

        # FEATURE 1: Only use active (checked) categories
        active_cats = self._get_active_categories()

        if not active_cats:
            messagebox.showwarning(
                "No categories",
                "Enable at least one category to run."
            )
            return

        do_csv = self.report_csv.get()
        do_excel = self.report_excel.get()

        report_categories = None
        if do_csv or do_excel:
            if self.report_cat_vars:
                report_categories = [
                    name for name, var in self.report_cat_vars.items()
                    if var.get() and name in active_cats
                ]
            else:
                report_categories = list(active_cats.keys())

            if not report_categories:
                messagebox.showwarning(
                    "Report",
                    "Report is enabled but no active categories selected."
                )
                return

        if do_excel and not HAS_OPENPYXL:
            messagebox.showwarning(
                "Missing Library",
                "openpyxl is required for Excel reports.\n"
                "Install with: pip install openpyxl"
            )
            return

        self._update_file_count()
        self.log_text.configure(state="normal")
        self.log_text.delete("1.0", "end")
        self.log_text.configure(state="disabled")
        self.run_btn.configure(state="disabled")
        self._is_running = True

        def on_done():
            self._is_running = False
            try:
                self.run_btn.configure(state="normal")
            except tk.TclError:
                pass

        def on_results(all_results, categories):
            """Open the report window on the main thread."""
            self.after(0, lambda: ReportWindow(
                self, all_results, categories, self.output_dir
            ))

        def task():
            try:
                process_all_pdfs(
                    self.input_dir, self.output_dir, active_cats,
                    log_func=lambda msg: self.after(0, self._log, msg),
                    done_callback=lambda: self.after(0, on_done),
                    report_categories=report_categories,
                    do_csv=do_csv,
                    do_excel=do_excel,
                    result_callback=on_results,  # FEATURE 2: triggers report window
                    stop_at_references=self.stop_at_refs.get()
                )
            except Exception as e:
                self.after(
                    0, self._log,
                    f"\n✖ Unexpected error: {e}"
                )
                self.after(0, on_done)

        threading.Thread(target=task, daemon=True).start()


if __name__ == "__main__":
    App().mainloop()
